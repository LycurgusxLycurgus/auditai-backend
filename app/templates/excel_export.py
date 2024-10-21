# app/templates/excel_export.py
import openpyxl
from openpyxl.utils import get_column_letter
from openpyxl.styles import Font
import os


def create_excel_report(json_data: dict, output_path: str):
    wb = openpyxl.Workbook()
    # Remove the default sheet
    default_sheet = wb.active
    wb.remove(default_sheet)

    consecutivo_contrato = json_data['contrato']['info_general'].get('Consecutivo Contrato', 'Unknown')

    # Sheet 1: Resumen Contratos
    resumen_sheet = wb.create_sheet(title=f"Resumen Contratos + {consecutivo_contrato}")
    resumen_headers = ["Campo", "Valor"]
    resumen_data = json_data['contrato']['info_general'].items()

    resumen_sheet.append(resumen_headers)
    for field, value in resumen_data:
        resumen_sheet.append([field, value])

    # Apply header styling
    for cell in resumen_sheet["1:1"]:
        cell.font = Font(bold=True)

    # Sheet 2: Verificación Clausulado
    clausulas_sheet = wb.create_sheet(title=f"Verificación Clausulado + {consecutivo_contrato}")
    clausulas_headers = ["Cláusula", "Existe", "Número", "Nombre", "Descripción"]
    clausulas_sheet.append(clausulas_headers)

    clausulas = json_data['contrato']['clausulas']['output']['Checklist de Cláusulas Contractuales']['Obligaciones de las Partes']
    for i in range(len(clausulas['Número de la cláusula'])):
        clausulas_sheet.append([
            "Obligaciones de las Partes",
            clausulas["Existe la cláusula"],
            clausulas["Número de la cláusula"][i],
            clausulas["Nombre de la cláusula"][i],
            clausulas["Descripción"][i],
        ])

    for cell in clausulas_sheet["1:1"]:
        cell.font = Font(bold=True)

    # Sheet 3: Pólizas
    poliza_sheet = wb.create_sheet(title=f"Pólizas + {consecutivo_contrato}")
    poliza_headers = ["Número Póliza", "Nombre Tomador", "Nombre Beneficiado", "Objeto", "Valor Total a Pagar", "Valor Total Asegurado", "Fecha de Pago"]
    poliza_sheet.append(poliza_headers)

    for poliza in json_data['poliza']:
        info = poliza['info_general']
        poliza_sheet.append([
            poliza['numero_poliza'],
            info.get('Nombre de tomador', 'NA'),
            info.get('Nombre de beneficiado', 'NA'),
            info.get('Objeto de la póliza', 'NA'),
            info.get('Valor total a pagar', 'NA'),
            info.get('Valor total asegurado', 'NA'),
            info.get('Fecha de pago', 'NA'),
        ])

    for cell in poliza_sheet["1:1"]:
        cell.font = Font(bold=True)

    # Sheet 4: Otrosíes
    otrosi_sheet = wb.create_sheet(title="Otrosíes")
    otrosi_headers = ["Número Otrosí", "Nit Tercero", "Nombre del Tercero", "Fecha Firma Otrosí", "Modificación"]
    otrosi_sheet.append(otrosi_headers)

    for otrosi in json_data['otrosi']:
        info_general = otrosi['info_general']
        modificaciones = otrosi['modificaciones']['output']['Modificaciones']['Valor_del_contrato']
        for mod in modificaciones:
            otrosi_sheet.append([
                otrosi['numero_otrosi'],
                info_general.get('Nit tercero', 'NA'),
                info_general.get('Nombre del Tercero', 'NA'),
                info_general.get('Fecha firma otrosi', 'NA'),
                mod.get('Modificacion_hecha', 'NA'),
            ])

    for cell in otrosi_sheet["1:1"]:
        cell.font = Font(bold=True)

    # Sheet 5: Hallazgos
    hallazgos_sheet = wb.create_sheet(title="Hallazgos")
    hallazgos_headers = ["Documento", "Campo", "Existe en Contrato", "Existe en Póliza", "Descripción"]
    hallazgos_sheet.append(hallazgos_headers)

    hallazgos = json_data.get('hallazgos', [])
    for hallazgo in hallazgos:
        for key, value in hallazgo['output']['0'].items():
            hallazgos_sheet.append([
                key,
                value.get('Existe en lista del contrato', 'NA'),
                value.get('Existe en documento póliza', 'NA'),
                value.get('Descripción', 'NA'),
            ])

    for cell in hallazgos_sheet["1:1"]:
        cell.font = Font(bold=True)

    # Auto-adjust column widths
    for sheet in wb.sheetnames:
        ws = wb[sheet]
        for column in ws.columns:
            max_length = 0
            column_letter = get_column_letter(column[0].column)
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = (max_length + 2)
            ws.column_dimensions[column_letter].width = adjusted_width

    # Save the workbook
    wb.save(output_path)
